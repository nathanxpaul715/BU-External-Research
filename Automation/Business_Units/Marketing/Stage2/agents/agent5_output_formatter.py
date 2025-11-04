"""Agent 5: Output Formatter & Excel Generator
Creates final Excel output with proper formatting
"""
import pandas as pd
from typing import Dict, List, Any
from datetime import datetime
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import OUTPUT_PATH, OUTPUT_COLUMNS


class OutputFormatterAgent:
    """Agent responsible for formatting and generating Excel output"""

    def __init__(self):
        self.output_data = []

    def format_cell_with_subheadings(self, text: str) -> str:
        """Format cell content with sub-headings and bullet points"""
        if not text:
            return ""

        # Ensure sub-headings are properly formatted
        # Replace sub-heading patterns with bold formatting (Excel will handle this)
        formatted = text.strip()

        # Add bullet points if not present
        lines = formatted.split('\n')
        formatted_lines = []
        for line in lines:
            line = line.strip()
            if line and not line.startswith('•') and not line.startswith('-'):
                # Check if it's a sub-heading (ends with colon)
                if ':' in line and len(line.split(':')[0].split()) <= 5:
                    formatted_lines.append(line)
                else:
                    formatted_lines.append(f"• {line}")
            elif line:
                formatted_lines.append(line)

        return '\n'.join(formatted_lines)

    def create_output_row(
        self,
        enriched_use_case: Dict[str, Any]
    ) -> Dict[str, str]:
        """Create a single output row for Excel"""

        original = enriched_use_case.get('original_use_case', {})
        enriched = enriched_use_case.get('enriched_data', {})

        row = {
            "Function": original.get('function', 'Marketing'),
            "Original Use Case Name": original.get('original_name', ''),
            "Enriched Use Case Name": enriched.get('enriched_name', original.get('original_name', '')),
            "Original Use Case Description": original.get('original_description', ''),
            "Detailed Enriched Use Case Description": self.format_cell_with_subheadings(
                enriched.get('detailed_description', '')
            ),
            "Original Outcomes/Deliverable": original.get('original_outcomes', ''),
            "Enriched Business Outcomes/Deliverables": self.format_cell_with_subheadings(
                enriched.get('business_outcomes', '')
            ),
            "Industry Alignment": self.format_cell_with_subheadings(
                enriched.get('industry_alignment', '')
            ),
            "Implementation Considerations": self.format_cell_with_subheadings(
                enriched.get('implementation', '')
            ),
            "Suggested Success Metrics (KPIs)": self.format_cell_with_subheadings(
                enriched.get('kpis', '')
            ),
            "Information Gaps & Annotation": self.format_cell_with_subheadings(
                enriched.get('annotation', '')
            )
        }

        return row

    def create_excel_output(self, enriched_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create Excel DataFrame with all enriched use cases"""
        print("Creating Excel output...")

        rows = []
        for enriched_use_case in enriched_data:
            if enriched_use_case.get('success', False):
                row = self.create_output_row(enriched_use_case)
                rows.append(row)
            else:
                # Handle failed enrichment
                original = enriched_use_case.get('original_use_case', {})
                row = {
                    "Function": original.get('function', 'Marketing'),
                    "Original Use Case Name": original.get('original_name', ''),
                    "Enriched Use Case Name": original.get('original_name', ''),
                    "Original Use Case Description": original.get('original_description', ''),
                    "Detailed Enriched Use Case Description": f"ERROR: {enriched_use_case.get('error', 'Unknown error')}",
                    "Original Outcomes/Deliverable": original.get('original_outcomes', ''),
                    "Enriched Business Outcomes/Deliverables": "",
                    "Industry Alignment": "",
                    "Implementation Considerations": "",
                    "Suggested Success Metrics (KPIs)": "",
                    "Information Gaps & Annotation": f"Error during enrichment: {enriched_use_case.get('error', 'Unknown error')}"
                }
                rows.append(row)

        df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
        print(f"✓ Created DataFrame with {len(df)} rows")

        return df

    def save_to_excel(self, df: pd.DataFrame) -> str:
        """Save DataFrame to Excel with formatting"""
        print(f"Saving to: {OUTPUT_PATH}")

        try:
            # Create Excel writer with openpyxl engine for formatting
            with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Enriched Use Cases', index=False)

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Enriched Use Cases']

                # Set column widths
                column_widths = {
                    'A': 15,  # Function
                    'B': 30,  # Original Use Case Name
                    'C': 30,  # Enriched Use Case Name
                    'D': 40,  # Original Use Case Description
                    'E': 60,  # Detailed Enriched Use Case Description
                    'F': 40,  # Original Outcomes/Deliverable
                    'G': 60,  # Enriched Business Outcomes/Deliverables
                    'H': 60,  # Industry Alignment
                    'I': 60,  # Implementation Considerations
                    'J': 60,  # Suggested Success Metrics (KPIs)
                    'K': 60   # Information Gaps & Annotation
                }

                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

                # Set row height and text wrapping
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = cell.alignment.copy(wrap_text=True, vertical='top')

                # Freeze header row
                worksheet.freeze_panes = 'A2'

                # Style header row
                from openpyxl.styles import Font, PatternFill
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font

            print(f"✓ Excel file saved successfully")
            return OUTPUT_PATH

        except Exception as e:
            print(f"✗ Error saving Excel: {e}")
            # Fallback to CSV
            csv_path = OUTPUT_PATH.replace('.xlsx', '.csv')
            df.to_csv(csv_path, index=False)
            print(f"✓ Saved as CSV instead: {csv_path}")
            return csv_path

    def run(self, enrichment_data: Dict[str, Any]) -> Dict[str, Any]:
        """Execute output formatting and Excel generation"""
        print("=" * 80)
        print("AGENT 5: OUTPUT FORMATTER & EXCEL GENERATOR")
        print("=" * 80)

        enriched_use_cases = enrichment_data.get("enriched_use_cases", [])

        # Create DataFrame
        df = self.create_excel_output(enriched_use_cases)

        # Save to Excel
        output_file = self.save_to_excel(df)

        print("\n✓ OUTPUT GENERATION COMPLETE")
        print("=" * 80)

        return {
            "output_file": output_file,
            "rows_written": len(df),
            "columns": list(df.columns),
            "timestamp": datetime.now().isoformat()
        }


if __name__ == "__main__":
    print("Output Formatter Agent - requires enriched data to run")
